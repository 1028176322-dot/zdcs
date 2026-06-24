#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
用例步骤解析器 - 按 AutoSmoke_按用例步骤自动点击方案.md 第5节实现

功能：
  将 Excel 用例中的步骤文本解析为结构化动作指令。

支持动作（文档 4.1 节）：
  点击、等待、断言存在、断言不存在、截图、返回

支持定位类型（文档 5.3 节）：
  text("xxx")      — OCR 文本定位
  template("xxx")  — 模板匹配定位
  normalized(x,y)  — 归一化坐标
  design(x,y)      — 设计分辨率坐标
  content(x,y)     — GameContent 内坐标
  testId("xxx")    — Unity 元数据定位（预留）

使用方式：
    parser = CaseStepParser()
    result = parser.parse('点击 text("使用")')
    # → {"action": "click", "target": {"type": "text", "value": "使用"}, "raw": "..."}
"""

import re
import json
import os
import logging
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ============================================================
# 动作关键词映射
# ============================================================

ACTION_KEYWORDS = {
    "点击": "click",
    "click": "click",
    "等待": "wait",
    "wait": "wait",
    "断言存在": "assert_exists",
    "断言不存在": "assert_not_exists",
    "截图": "screenshot",
    "screenshot": "screenshot",
    "返回": "back",
    "back": "back",
    "长按": "long_press",
    "滑动": "swipe",
    "swipe": "swipe",
    "输入": "input",
}


class CaseStepParser:
    """
    用例步骤解析器

    将文本步骤解析为结构化动作指令，支持中文和英文关键词。
    """

    def __init__(self):
        """初始化解析器"""
        self._action_pattern = self._build_action_pattern()

    @staticmethod
    def _build_action_pattern() -> re.Pattern:
        """构建动作匹配正则"""
        # 按长度降序排列（避免"断言存在"被"断言"先匹配）
        actions = sorted(ACTION_KEYWORDS.keys(), key=len, reverse=True)
        pattern = r"^(?P<action>" + "|".join(re.escape(a) for a in actions) + r")\s*"
        return re.compile(pattern)

    # ============================================================
    # 定位类型解析
    # ============================================================

    # 各定位类型的正则（支持 " 双引号、' 单引号、"" 中文引号）
    _TEXT_PATTERN = re.compile(
        r"""text\s*\(\s*["'\u201C\u201D](?P<value>[^"'\u201D]*)["'\u201D]\s*\)"""
    )
    _TEMPLATE_PATTERN = re.compile(
        r"""template\s*\(\s*["'\u201C\u201D](?P<value>[^"'\u201D]*)["'\u201D]\s*\)"""
    )
    _NORMALIZED_PATTERN = re.compile(
        r'normalized\s*\(\s*(?P<nx>[\d.]+)\s*,\s*(?P<ny>[\d.]+)\s*\)'
    )
    _DESIGN_PATTERN = re.compile(
        r'design\s*\(\s*(?P<x>[\d.]+)\s*,\s*(?P<y>[\d.]+)\s*\)'
    )
    _CONTENT_PATTERN = re.compile(
        r'content\s*\(\s*(?P<x>[\d.]+)\s*,\s*(?P<y>[\d.]+)\s*\)'
    )
    _TESTID_PATTERN = re.compile(
        r"""testId\s*\(\s*["'\u201C\u201D](?P<value>[^"'\u201D]*)["'\u201D]\s*\)"""
    )
    _PIXEL_PATTERN = re.compile(
        r'pixel\s*\(\s*(?P<x>[\d.]+)\s*,\s*(?P<y>[\d.]+)\s*\)'
    )

    # 按优先级排序的定位解析器列表
    _LOCATORS = [
        ("text", _TEXT_PATTERN),
        ("template", _TEMPLATE_PATTERN),
        ("testId", _TESTID_PATTERN),
        ("normalized", _NORMALIZED_PATTERN),
        ("design", _DESIGN_PATTERN),
        ("content", _CONTENT_PATTERN),
        ("pixel", _PIXEL_PATTERN),
    ]

    def _parse_target(self, step_text: str) -> Optional[Dict]:
        """
        从步骤文本中解析定位目标

        :param step_text: 去除动作前缀后的步骤文本
        :return: {"type": "text", "value": "使用"} 或 None
        """
        for loc_type, pattern in self._LOCATORS:
            match = pattern.search(step_text)
            if match:
                groups = match.groupdict()
                if loc_type in ("normalized",):
                    return {
                        "type": loc_type,
                        "nx": float(groups["nx"]),
                        "ny": float(groups["ny"]),
                    }
                elif loc_type in ("design", "content", "pixel"):
                    return {
                        "type": loc_type,
                        "x": float(groups["x"]),
                        "y": float(groups["y"]),
                    }
                else:
                    # text, template, testId
                    return {
                        "type": loc_type,
                        "value": groups["value"].strip(),
                    }
        return None

    # ============================================================
    # 等待时长解析
    # ============================================================

    _WAIT_SEC_PATTERN = re.compile(r"(\d+\.?\d*)\s*秒")
    _WAIT_S_PATTERN = re.compile(r"(\d+\.?\d*)\s*s", re.IGNORECASE)

    def _parse_wait_seconds(self, step_text: str) -> Optional[float]:
        """从等待步骤中解析秒数"""
        for pat in (self._WAIT_SEC_PATTERN, self._WAIT_S_PATTERN):
            match = pat.search(step_text)
            if match:
                return float(match.group(1))
        # 没有显式秒数，默认 2 秒
        return 2.0

    # ============================================================
    # 主解析方法
    # ============================================================

    def parse(self, step_text: str) -> Dict:
        """
        解析单条步骤文本

        :param step_text: 步骤文本，如 '点击 text("使用")'
        :return:
            {
                "action": "click",
                "target": {"type": "text", "value": "使用"},
                "raw": "点击 text(\"使用\")",
                "valid": True
            }
            解析失败时：
            {
                "action": "unknown",
                "raw": "...",
                "valid": False,
                "error": "无法识别动作"
            }
        """
        step_text = step_text.strip()
        if not step_text:
            return {"action": "empty", "raw": "", "valid": False, "error": "空步骤"}

        # 去除注释（中文/英文分号后的内容）
        clean_text = re.split(r"[;#；]", step_text)[0].strip()

        # 匹配动作
        action_match = self._action_pattern.match(clean_text)
        if not action_match:
            return {
                "action": "unknown",
                "raw": step_text,
                "valid": False,
                "error": f"无法识别动作: {clean_text[:30]}",
            }

        action = ACTION_KEYWORDS[action_match.group("action")]
        rest = clean_text[action_match.end():].strip()

        result = {
            "action": action,
            "raw": step_text,
            "valid": True,
        }

        # 按动作类型解析
        if action in ("click", "long_press"):
            target = self._parse_target(rest)
            if target:
                result["target"] = target
            else:
                result["valid"] = False
                result["error"] = f"点击步骤缺少有效定位: {rest[:30]}"

        elif action == "wait":
            result["seconds"] = self._parse_wait_seconds(rest)

        elif action in ("assert_exists", "assert_not_exists"):
            target = self._parse_target(rest)
            if target:
                result["target"] = target
            else:
                result["valid"] = False
                result["error"] = f"断言步骤缺少有效定位: {rest[:30]}"

        elif action == "input":
            # 支持: 输入 text("xxx") 到 text("field")
            target = self._parse_target(rest)
            if target:
                result["target"] = target
            # 如果还有 value，提取
            value_match = re.search(r'["\u201C\u201D]([^"\u201D]*)["\u201D]', rest)
            if value_match:
                result["value"] = value_match.group(1)

        elif action == "swipe":
            # 滑动格式: 滑动 (x1,y1,x2,y2) 或 normalized(x1,y1,x2,y2)
            nums = re.findall(r"[\d.]+", rest)
            if len(nums) >= 4:
                result["start"] = (float(nums[0]), float(nums[1]))
                result["end"] = (float(nums[2]), float(nums[3]))

        elif action in ("screenshot", "back"):
            # 无需额外参数
            pass

        return result

    # ============================================================
    # 批量解析
    # ============================================================

    def parse_all(self, steps: List[str]) -> List[Dict]:
        """
        批量解析多条步骤

        :param steps: 步骤文本列表
        :return: 解析结果列表
        """
        return [self.parse(s) for s in steps]

    def parse_from_excel(self, rows: List[Dict],
                         text_field: str = "操作步骤") -> List[Dict]:
        """
        从 Excel 行数据解析步骤

        :param rows: Excel 行字典列表
        :param text_field: 步骤文本的列名
        :return: 解析结果列表
        """
        results = []
        for row in rows:
            step_text = row.get(text_field, "")
            if step_text:
                parsed = self.parse(step_text)
                # 保留 Excel 中的其他字段
                parsed["_excel_row"] = {k: v for k, v in row.items()
                                        if k != text_field}
                results.append(parsed)
        return results

    def parse_from_excel_file(self, filepath: str,
                              sheet_name: str = None,
                              text_field: str = "操作步骤",
                              header_row: int = None) -> List[Dict]:
        """
        从 .xlsx 文件直接读取并解析步骤

        :param filepath: .xlsx 文件路径
        :param sheet_name: 工作表名称（默认第一个非空sheet）
        :param text_field: 步骤文本的列名（默认"操作步骤"）
        :param header_row: 表头行号（从1开始，默认自动检测）
        :return: 解析结果列表
        :raises FileNotFoundError: 文件不存在
        :raises ValueError: 列名不匹配或无数据
        """
        import openpyxl

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)

        # 选择工作表
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # 自动检测表头行
        if header_row is None:
            header_row = self._detect_header_row(ws, text_field)
            if header_row is None:
                raise ValueError(
                    f"未找到包含 '{text_field}' 列的表头行")

        # 读取表头
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value).strip()
                           if cell.value is not None else "")

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1,
                                max_row=ws.max_row, values_only=True):
            row_dict = {}
            has_data = False
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    key = headers[i]
                    if val is not None:
                        row_dict[key] = str(val).strip()
                        has_data = True
                    else:
                        row_dict[key] = ""
            if has_data:
                rows.append(row_dict)

        wb.close()

        if not rows:
            raise ValueError("Excel 文件中没有数据行")

        logger.info("Excel 读取完成: %s → %d 行 (表头行=%d, text_field=%s)",
                     filepath, len(rows), header_row, text_field)

        # 复用现有解析逻辑
        return self.parse_from_excel(rows, text_field)

    def _detect_header_row(self, ws, text_field: str) -> int:
        """自动检测包含 text_field 的表头行"""
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            for cell in ws[row_idx]:
                if cell.value and str(cell.value).strip() == text_field:
                    return row_idx
        return None

    # ============================================================
    # 工具方法
    # ============================================================

    def to_json(self, parsed: Dict, indent: int = 2) -> str:
        """将解析结果格式化为 JSON 字符串"""
        return json.dumps(parsed, indent=indent, ensure_ascii=False)

    def validate_step(self, step_text: str) -> Dict:
        """
        验证步骤文本是否有效，返回详细校验结果

        :param step_text: 步骤文本
        :return: {"valid": True/False, "parsed": {...}, "warnings": [...]}
        """
        result = self.parse(step_text)
        warnings = []

        if not result["valid"]:
            return {"valid": False, "parsed": result,
                    "warnings": [result.get("error", "未知错误")]}

        # 额外校验
        target = result.get("target", {})
        if result["action"] == "click" and not target:
            warnings.append("点击步骤缺少目标")

        if target and target.get("type") == "text" and not target.get("value"):
            warnings.append("text 定位值为空")

        if target and target.get("type") == "template" and not target.get("value"):
            warnings.append("template 定位值为空")

        return {"valid": len(warnings) == 0, "parsed": result,
                "warnings": warnings}

    def summarize(self, parsed_list: List[Dict]) -> Dict:
        """
        汇总多条解析结果

        :param parsed_list: 解析结果列表
        :return: {"total": N, "valid": N, "invalid": N, "actions": {...}}
        """
        total = len(parsed_list)
        valid = sum(1 for p in parsed_list if p.get("valid"))
        invalid = total - valid
        action_counts = {}
        for p in parsed_list:
            a = p.get("action", "unknown")
            action_counts[a] = action_counts.get(a, 0) + 1

        return {
            "total": total,
            "valid": valid,
            "invalid": invalid,
            "actions": action_counts,
        }


# ============================================================
# 独立运行入口
# ============================================================

def test_parser():
    """测试解析器"""
    print("=" * 60)
    print("用例步骤解析器测试")
    print("=" * 60)

    parser = CaseStepParser()

    # ========== 测试用例集 ==========
    test_cases = [
        # (输入, 期望action, 期望target_type, 期望target_value/key)
        ("点击 text(\"使用\")", "click", "text", {"value": "使用"}),
        ("点击 text('金币')", "click", "text", {"value": "金币"}),
        ("点击 normalized(0.5,0.95)", "click", "normalized", {"nx": 0.5, "ny": 0.95}),
        ("点击 design(585,2400)", "click", "design", {"x": 585, "y": 2400}),
        ("点击 content(160,665)", "click", "content", {"x": 160, "y": 665}),
        ("点击 template(\"use_button\")", "click", "template", {"value": "use_button"}),
        ("点击 pixel(100,200)", "click", "pixel", {"x": 100, "y": 200}),
        ("等待 2 秒", "wait", None, {"seconds": 2.0}),
        ("等待 5s", "wait", None, {"seconds": 5.0}),
        ("断言存在 text(\"联盟迁城\")", "assert_exists", "text", {"value": "联盟迁城"}),
        ("断言不存在 text(\"加载中\")", "assert_not_exists", "text", {"value": "加载中"}),
        ("截图", "screenshot", None, None),
        ("返回", "back", None, None),
        ("click design(585,2400)", "click", "design", {"x": 585, "y": 2400}),
        ("wait 3 s", "wait", None, {"seconds": 3.0}),
        # 无效用例
        ("随便写点什么", "unknown", None, None),
        ("", "empty", None, None),
    ]

    passed = 0
    failed = 0

    for i, (input_text, exp_action, exp_type, exp_values) in enumerate(test_cases, 1):
        result = parser.parse(input_text)
        action_ok = result["action"] == exp_action
        type_ok = True
        values_ok = True

        if exp_type is not None:
            target = result.get("target", {})
            type_ok = target.get("type") == exp_type
            if exp_values and type_ok:
                for k, v in exp_values.items():
                    actual = target.get(k)
                    if isinstance(v, float):
                        values_ok = values_ok and abs(actual - v) < 0.001
                    else:
                        values_ok = values_ok and actual == v

        all_ok = action_ok and type_ok and values_ok
        if all_ok:
            passed += 1
            status = "✅"
        else:
            failed += 1
            status = "❌"

        print(f"\n[{status} 测试{i}] {input_text}")
        print(f"  → {json.dumps(result, ensure_ascii=False)}")

    # ========== 批量解析测试 ==========
    print(f"\n{'=' * 60}")
    print("[批量解析测试]")
    steps = [
        "点击 text(\"使用\")",
        "等待 2 秒",
        "断言存在 text(\"联盟迁城\")",
        "截图",
        "返回",
    ]
    results = parser.parse_all(steps)
    s = parser.summarize(results)
    print(f"  总数: {s['total']}, 有效: {s['valid']}, 无效: {s['invalid']}")
    print(f"  动作分布: {s['actions']}")
    if s["invalid"] == 0:
        passed += 1
        print("  ✅ 批量解析通过")
    else:
        failed += 1
        print("  ❌ 批量解析失败")

    # ========== 校验测试 ==========
    print(f"\n{'=' * 60}")
    print("[校验测试]")
    v = parser.validate_step("点击 text(\"\")")
    print(f"  空文本: valid={v['valid']}, warnings={v['warnings']}")
    v2 = parser.validate_step("点击 normalized(0.5,0.95)")
    print(f"  正常步骤: valid={v2['valid']}, warnings={v2['warnings']}")
    passed += 1

    # ========== 汇总 ==========
    print(f"\n{'=' * 60}")
    print(f"通过: {passed} / 失败: {failed}")
    if failed == 0:
        print("所有测试通过 ✅")
    else:
        print("有测试失败 ❌")


if __name__ == "__main__":
    test_parser()
