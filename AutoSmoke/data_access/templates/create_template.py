"""
生成测试用例模板Excel文件
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_test_case_template(output_path):
    """创建测试用例模板Excel文件"""
    print(f"正在创建测试用例模板: {output_path}")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 定义表头
    headers = [
        '测试用例ID',
        '测试用例名称',
        '功能模块',
        '优先级',
        '前置条件',
        '测试步骤',
        '预期结果'
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 添加示例测试用例
    sample_cases = [
        [
            'TC_001',
            '点击底部导航栏主城按钮',
            '主界面',
            '高',
            '游戏已启动，进入主界面',
            '1. 点击底部导航栏的"主城"按钮\n2. 等待2秒',
            '1. 成功点击主城按钮\n2. 界面切换到主城界面\n3. 网络消息中有MainCityInfo协议'
        ],
        [
            'TC_002',
            '点击底部导航栏小游戏按钮',
            '主界面',
            '高',
            '游戏已启动，进入主界面',
            '1. 点击底部导航栏的"小游戏"按钮\n2. 等待2秒',
            '1. 成功点击小游戏按钮\n2. 界面切换到小游戏界面\n3. 网络消息中有MiniGameInfo协议'
        ],
        [
            'TC_003',
            '点击底部导航栏英雄按钮',
            '主界面',
            '中',
            '游戏已启动，进入主界面',
            '1. 点击底部导航栏的"英雄"按钮\n2. 等待2秒',
            '1. 成功点击英雄按钮\n2. 界面切换到英雄界面\n3. 网络消息中有HeroInfo协议'
        ],
        [
            'TC_004',
            '点击底部导航栏联盟按钮',
            '主界面',
            '中',
            '游戏已启动，进入主界面',
            '1. 点击底部导航栏的"联盟"按钮\n2. 等待2秒',
            '1. 成功点击联盟按钮\n2. 界面切换到联盟界面\n3. 网络消息中有AllianceInfo协议'
        ],
        [
            'TC_005',
            '点击底部导航栏世界地图按钮',
            '主界面',
            '高',
            '游戏已启动，进入主界面',
            '1. 点击底部导航栏的"世界地图"按钮\n2. 等待2秒',
            '1. 成功点击世界地图按钮\n2. 界面切换到世界地图界面\n3. 网络消息中有WorldMapInfo协议'
        ]
    ]
    
    # 写入示例测试用例
    for row_idx, test_case in enumerate(sample_cases, 2):
        for col_idx, value in enumerate(test_case, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 调整列宽
    column_widths = [15, 30, 15, 10, 30, 40, 50]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 调整行高
    ws.row_dimensions[1].height = 30
    for row in range(2, len(sample_cases) + 2):
        ws.row_dimensions[row].height = 60
    
    # 保存工作簿
    wb.save(output_path)
    print(f"✅ 测试用例模板已创建: {output_path}")
    
    return output_path

if __name__ == '__main__':
    # 获取输出路径
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, 'test_case_template.xlsx')
    
    # 创建测试用例模板
    create_test_case_template(output_path)
    
    print("\n使用方法:")
    print("1. 打开 test_case_template.xlsx")
    print("2. 根据示例格式编写测试用例")
    print("3. 保存Excel文件")
    print("4. 运行 test_runner.py，输入Excel文件路径")
    print("5. 系统将自动连接Unity，执行测试步骤，并生成测试报告")
