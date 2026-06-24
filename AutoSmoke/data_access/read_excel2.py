import sys
import os
from pathlib import Path

file = str((Path(__file__).resolve().parents[1] / "参考资料" / "用例模板.xlsx"))
print(f"尝试读取文件: {file}")
print(f"文件是否存在: {os.path.exists(file)}")

try:
    # 尝试使用openpyxl
    import openpyxl
    print("使用 openpyxl 读取...")
    
    wb = openpyxl.load_workbook(file, data_only=True)
    print(f'\n工作表列表: {wb.sheetnames}')
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f'\n====== 工作表: {sheet_name} ======')
        print(f'最大行数: {sheet.max_row}, 最大列数: {sheet.max_column}')
        
        # 读取前10行数据
        print('\n数据内容:')
        for row in sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=True):
            if any(cell is not None for cell in row):
                print(row)
                
except ImportError:
    print("openpyxl 未安装，尝试使用pandas...")
    try:
        import pandas as pd
        xl = pd.ExcelFile(file)
        print(f'\n工作表列表: {xl.sheet_names}')
        
        for sheet in xl.sheet_names:
            df = pd.read_excel(file, sheet_name=sheet)
            print(f'\n====== 工作表: {sheet} ======')
            print(f'行数: {len(df)}, 列数: {len(df.columns)}')
            print(f'列名: {list(df.columns)}')
            print('\n所有数据:')
            print(df.to_string())
    except ImportError:
        print("pandas 也未安装，无法读取Excel文件")
        sys.exit(1)
except Exception as e:
    print(f'错误: {e}')
    import traceback
    traceback.print_exc()
