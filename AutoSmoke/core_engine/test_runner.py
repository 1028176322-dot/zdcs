"""
# -*- coding: utf-8 -*-
自动化测试执行系统 - 主程序
读取Excel测试用例，连接Unity，执行测试步骤，生成测试报告
"""

import os
import sys
import time
import json
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

# 导入自定义模块
from test_executor import TestExecutor
from test_reporter import TestReporter
from config import Config

class TestRunner:
    def __init__(self, config_path=None):
        self.config = Config(config_path)
        self.executor = None
        self.reporter = None
        self.test_cases = []
        self.results = []
        
    def load_test_cases(self, excel_path):
        """从Excel文件加载测试用例"""
        print(f"正在读取测试用例: {excel_path}")
        
        try:
            # 使用pandas读取Excel
            df = pd.read_excel(excel_path)
            
            # 转换为测试用例列表
            for index, row in df.iterrows():
                test_case = {
                    'id': row.get('测试用例ID', f'TC_{index+1:03d}'),
                    'name': row.get('测试用例名称', f'测试用例{index+1}'),
                    'precondition': row.get('前置条件', ''),
                    'steps': self._parse_steps(row.get('测试步骤', '')),
                    'expected': row.get('预期结果', ''),
                    'priority': row.get('优先级', '中'),
                    'module': row.get('功能模块', '未分类')
                }
                self.test_cases.append(test_case)
            
            print(f"成功加载 {len(self.test_cases)} 个测试用例")
            return True
            
        except Exception as e:
            print(f"读取测试用例失败: {e}")
            return False
    
    def _parse_steps(self, steps_text):
        """解析测试步骤文本（支持多种格式）"""
        if pd.isna(steps_text):
            return []
        
        steps = []
        # 尝试按行分割
        lines = str(steps_text).split('\n')
        for i, line in enumerate(lines, 1):
            line = line.strip()
            if line:
                # 移除步骤编号（如 "1. "、"1、"等）
                import re
                line = re.sub(r'^\d+[\.\、\)\]]\s*', '', line)
                steps.append({
                    'step_id': i,
                    'action': line,
                    'type': self._detect_step_type(line)
                })
        
        return steps
    
    def _detect_step_type(self, action_text):
        """检测步骤类型"""
        action_text = action_text.lower()
        
        if any(keyword in action_text for keyword in ['点击', 'click', '触摸', 'tap']):
            return 'click'
        elif any(keyword in action_text for keyword in ['输入', '填写', '输入文本', 'type', 'enter']):
            return 'input'
        elif any(keyword in action_text for keyword in ['等待', 'sleep', '延时', '延迟']):
            return 'wait'
        elif any(keyword in action_text for keyword in ['检查', '验证', '断言', 'assert', 'check', 'verify']):
            return 'verify'
        elif any(keyword in action_text for keyword in ['截图', 'screenshot', '捕获']):
            return 'screenshot'
        else:
            return 'unknown'
    
    def connect_unity(self):
        """连接Unity游戏"""
        print("正在连接Unity游戏...")
        
        try:
            self.executor = TestExecutor(self.config)
            success = self.executor.connect()
            
            if success:
                print("✅ Unity连接成功")
                return True
            else:
                print("❌ Unity连接失败")
                return False
                
        except Exception as e:
            print(f"连接Unity时出错: {e}")
            return False
    
    def run_all_tests(self):
        """运行所有测试用例"""
        if not self.test_cases:
            print("没有测试用例可执行")
            return False
        
        print(f"\n开始执行 {len(self.test_cases)} 个测试用例")
        print("="*50)
        
        # 初始化报告器
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_dir = os.path.join(self.config.report_dir, timestamp)
        os.makedirs(report_dir, exist_ok=True)
        self.reporter = TestReporter(report_dir)
        
        # 执行每个测试用例
        for i, test_case in enumerate(self.test_cases, 1):
            print(f"\n[{i}/{len(self.test_cases)}] 执行: {test_case['name']}")
            
            result = self._execute_test_case(test_case)
            self.results.append(result)
            
            # 实时更新报告
            self.reporter.update_result(result)
            
            # 打印结果
            status = "✅ 通过" if result['status'] == 'PASS' else "❌ 失败"
            print(f"结果: {status} - {result['message']}")
        
        # 生成最终报告
        report_path = self.reporter.generate_report(self.results)
        print(f"\n测试完成！报告已生成: {report_path}")
        
        return True
    
    def _execute_test_case(self, test_case):
        """执行单个测试用例"""
        result = {
            'id': test_case['id'],
            'name': test_case['name'],
            'module': test_case['module'],
            'priority': test_case['priority'],
            'status': 'FAIL',
            'message': '',
            'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'end_time': '',
            'duration': 0,
            'steps_executed': 0,
            'steps_total': len(test_case['steps']),
            'screenshots': [],
            'protocols': []
        }
        
        start_time = time.time()
        
        try:
            # 检查前置条件
            if test_case['precondition']:
                print(f"  前置条件: {test_case['precondition']}")
                # TODO: 实现前置条件检查
            
            # 执行测试步骤
            for step in test_case['steps']:
                print(f"  步骤{step['step_id']}: {step['action']}")
                
                step_result = self.executor.execute_step(step)
                
                result['steps_executed'] += 1
                
                if not step_result['success']:
                    result['status'] = 'FAIL'
                    result['message'] = f"步骤{step['step_id']}失败: {step_result['message']}"
                    break
                
                # 收集协议信息
                if 'protocols' in step_result:
                    result['protocols'].extend(step_result['protocols'])
                
                # 收集截图信息
                if 'screenshot' in step_result:
                    result['screenshots'].append(step_result['screenshot'])
            
            # 如果所有步骤都成功执行，则标记为通过
            if result['steps_executed'] == result['steps_total']:
                result['status'] = 'PASS'
                result['message'] = '所有步骤执行成功'
                
                # 验证预期结果
                if test_case['expected']:
                    print(f"  验证预期结果: {test_case['expected']}")
                    # TODO: 实现预期结果验证
                    
        except Exception as e:
            result['status'] = 'ERROR'
            result['message'] = f"执行出错: {str(e)}"
        
        # 计算执行时间
        end_time = time.time()
        result['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result['duration'] = round(end_time - start_time, 2)
        
        return result
    
    def save_results(self, output_path=None):
        """保存测试结果到Excel"""
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(self.config.report_dir, f'test_results_{timestamp}.xlsx')
        
        print(f"正在保存测试结果: {output_path}")
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "测试结果"
            
            # 写入表头
            headers = ['测试用例ID', '测试用例名称', '功能模块', '优先级', '测试状态', '消息', '开始时间', '结束时间', '执行时长(秒)', '执行步骤数', '总步骤数']
            ws.append(headers)
            
            # 设置表头样式
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # 写入测试结果
            for result in self.results:
                row = [
                    result['id'],
                    result['name'],
                    result['module'],
                    result['priority'],
                    result['status'],
                    result['message'],
                    result['start_time'],
                    result['end_time'],
                    result['duration'],
                    result['steps_executed'],
                    result['steps_total']
                ]
                ws.append(row)
                
                # 设置状态单元格颜色
                status_cell = ws.cell(row=ws.max_row, column=5)
                if result['status'] == 'PASS':
                    status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif result['status'] == 'FAIL':
                    status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存工作簿
            wb.save(output_path)
            print(f"✅ 测试结果已保存: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ 保存测试结果失败: {e}")
            return False

def main():
    """主函数"""
    print("="*50)
    print("自动化测试执行系统")
    print("="*50)
    
    # 创建测试运行器
    runner = TestRunner()
    
    # 读取测试用例
    excel_path = input("请输入测试用例Excel文件路径: ").strip()
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return
    
    if not runner.load_test_cases(excel_path):
        return
    
    # 连接Unity
    if not runner.connect_unity():
        print("无法连接Unity，请确认游戏已启动")
        return
    
    # 运行测试
    runner.run_all_tests()
    
    # 保存结果
    runner.save_results()
    
    print("\n测试执行完成！")

if __name__ == '__main__':
    main()
